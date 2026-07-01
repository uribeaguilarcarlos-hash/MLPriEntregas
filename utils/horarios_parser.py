import openpyxl

def parsear_horarios(wb):
    ws = wb.active
    horarios = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[4]:
            continue
        cont = str(row[4]).strip()
        if cont and cont not in horarios:
            horarios[cont] = {
                "fecha": str(row[1]) if row[1] else "",
                "hora":  str(row[2]) if row[2] else "",
                "cortina": str(row[5]) if row[5] else ""
            }
    return horarios

def actualizar_horarios_excel(ws, filas_nuevas: list) -> int:
    """
    Reemplaza filas del mismo día o agrega si no existen.
    Detecta el día por la fecha de la primera fila nueva.
    """
    if not filas_nuevas:
        return 0

    fecha_nueva = filas_nuevas[0].get("fecha", "")
    
    # Encontrar y borrar filas existentes del mismo día
    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[1] and fecha_nueva and str(row[1]).startswith(fecha_nueva[:10]):
            rows_to_delete.append(i)
    
    # Borrar en reversa para no correr índices
    for i in reversed(rows_to_delete):
        ws.delete_rows(i)
    
    # Agregar nuevas filas al final
    last = ws.max_row
    for fila in filas_nuevas:
        last += 1
        ws.cell(last, 1, fila.get("numero", ""))
        ws.cell(last, 2, fila.get("fecha", ""))
        ws.cell(last, 3, fila.get("hora", ""))
        ws.cell(last, 4, fila.get("rubro", ""))
        ws.cell(last, 5, fila.get("contenedor", ""))
        ws.cell(last, 6, fila.get("cortina", ""))
    
    return len(filas_nuevas)